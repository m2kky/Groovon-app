"""
Process David's Excel file: extract artists, classify, enrich, fill all columns.

Pipeline:
1. Read Excel events (city, venue, event_string)
2. AI batch classify: extract artist names, genre, event type
2.5. Search full lineups for "+more" events
3. Spotify verify + grab photo/genres/url
4. AI enrich: bio, locale, URLs, emails (3-pass search)
5. Write completed Excel

Usage:
  python process_david_excel.py             # Process all events
  python process_david_excel.py 210         # Start from event 210 (resume)
  python process_david_excel.py 0 100       # Process first 100 events

API Priority: Google AI Studio (free) → OpenRouter (paid)
"""
import os, json, urllib.request, urllib.parse, time, logging, re, sys, base64, threading, socket, html
from pipeline.checkpoint import Checkpoint
from pipeline.config import (
    SUPABASE_URL, SUPABASE_KEY, OPENROUTER_KEY, GOOGLE_API_KEY,
    SPOTIFY_ID, SPOTIFY_SECRET, LASTFM_API_KEY, YOUTUBE_API_KEY,
    DISCOGS_TOKEN, GENIUS_TOKEN, GOOGLE_KG_KEY, SERPER_KEY,
    SETLISTFM_KEY, SCRAPINGBEE_KEY,
    SSL_CTX, GENRES, validate_genre, best_genre,
    VALID_TYPES, DELETE_TYPES, sanitize_event_type,
    classify_url, MB_LINK_MAP, normalize,
)
from pipeline.ai_engine import (
    ai_call, parse_ai_result, classify_batch, enrich_batch,
    search_event_lineup,
)
from concurrent.futures import ThreadPoolExecutor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)-6s %(message)s')
log = logging.getLogger(__name__)

from pipeline.fetchers import (
    spotify_search, verify_artist_multiplatform,
    mb_get_links, lastfm_search, wikipedia_search, wikidata_search,
    itunes_search, discogs_search, youtube_search, genius_search,
    google_kg_search, setlistfm_search,
    bandsintown_search, soundcloud_search, bandcamp_search,
    deezer_extended_search, extract_mb_extra_links,
    build_profile, scrape_official_site, scrape_linktree,
    serper_search, ddg_search,
    verify_url, validate_email,
    parallel_fetch,
)
from pipeline.excel_io import (
    write_output, HEADER_ROW, calc_confidence,
    validate_inputs, calc_profile_score,
)
from pipeline.validator import rank_emails, email_label, validate_output_report
from pipeline.supabase_uploader import upload_profiles

def parse_args():
    """Parse command-line arguments for the pipeline."""
    import argparse
    parser = argparse.ArgumentParser(
        description="Process David's Excel: classify, enrich, and fill artist data.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""\
examples:
  python process_david_excel.py                                  # Process all events (defaults)
  python process_david_excel.py --start 210                      # Resume from event 210
  python process_david_excel.py --start 0 --limit 100            # First 100 events
  python process_david_excel.py --input my_file.xlsx --output out.xlsx
  python process_david_excel.py --dry-run                        # Preview only, no writes
  python process_david_excel.py --mode artists --input artists.xlsx   # Pre-extracted artist list
""",
    )
    parser.add_argument("--input", "-i", default=r"d:\projects\groovon\new\practise file for Makky Feb.xlsx",
                        help="Path to input Excel file (default: practise file for Makky Feb.xlsx)")
    parser.add_argument("--output", "-o", default=r"d:\projects\groovon\new\practise file COMPLETED.xlsx",
                        help="Path to output Excel file (default: practise file COMPLETED.xlsx)")
    parser.add_argument("--start", "-s", type=int, default=0,
                        help="Start processing from this event index (default: 0)")
    parser.add_argument("--limit", "-l", type=int, default=9999,
                        help="Max number of events to process (default: 9999)")
    parser.add_argument("--batch-size", "-b", type=int, default=15,
                        help="AI classification batch size (default: 15)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview mode: classify and display results without writing output")
    parser.add_argument("--mode", choices=["events", "artists"], default="events",
                        help="Input mode: 'events' (default, classify+enrich) or 'artists' (pre-extracted, skip classify)")
    parser.add_argument("--sheet", default=None,
                        help="Sheet name to read (artist mode only; auto-detects if omitted)")
    return parser.parse_args()


def main():
    args = parse_args()
    INPUT_FILE  = args.input
    OUTPUT_FILE = args.output
    batch_size  = args.batch_size
    start_from  = args.start
    limit       = args.limit
    dry_run     = args.dry_run

    # ── Artist-list mode: delegate to GroovonEngine directly ──
    if args.mode == "artists":
        from sources.artist_list_source import ArtistListSource
        from sinks.excel_sink import ExcelSink
        from engine import GroovonEngine

        source = ArtistListSource(config={
            "path": INPUT_FILE,
            "sheet": args.sheet,
        })
        sinks = [ExcelSink(config={"output_path": OUTPUT_FILE})]
        engine = GroovonEngine(source=source, sinks=sinks, dry_run=dry_run)
        engine.run(batch_size=batch_size, start=start_from, limit=limit)
        return
    
    # ── Step 0: Validate input (Item 9) ──
    try:
        validate_inputs(INPUT_FILE)
    except ValueError as exc:
        log.error(f"❌ Input validation failed: {exc}")
        return

    # ── Step 1: Read Excel ──
    log.info("📖 Reading Excel file...")
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb['base info']
    
    events = []
    for r in range(4, ws.max_row + 1):
        city = ws.cell(r, 2).value
        venue = ws.cell(r, 3).value
        title = ws.cell(r, 4).value
        if not title:
            continue
        events.append({"row": r, "city": str(city or "").strip(), "venue": str(venue or "").strip(), "title": str(title).strip()})
    
    log.info(f"   {len(events)} events read")
    events = events[start_from:start_from + limit]
    log.info(f"   Processing {len(events)} events (start={start_from}, limit={limit})\n")
    
    # ── Checkpoint — resume from last completed phase ──
    batch_id = os.path.splitext(os.path.basename(INPUT_FILE))[0].replace(' ', '_')
    cp = Checkpoint(batch_id)
    _cp_state = cp.load()
    _cp_phase = _cp_state["phase"] if _cp_state else 0
    _cp_data  = _cp_state["data"]  if _cp_state else {}
    if _cp_phase:
        log.info(f"   ⏩ Resuming — skipping phases already completed (up to phase {_cp_phase})\n")
    
    # Test Spotify — auto-disable if rate-limited
    log.info("🎵 Testing Spotify...")
    test = spotify_search("Gorillaz")
    if test:
        log.info("   ✅ OK\n")
    else:
        _spotify_disabled = True
        log.info("   ⚠️ Failed — Spotify disabled for this run\n")
    
    # ── Step 2: AI Classification (batches) ──
    if _cp_phase >= 1:
        classified = _cp_data["classified"]
        log.info(f"🤖 Phase 1: ⏩ Restored {len(classified)} classified events from checkpoint")
    else:
        log.info("🤖 Phase 1: AI Classification (David's methodology)...")
        classified = []
        total_batches = (len(events) + batch_size - 1) // batch_size
        
        for batch_start in range(0, len(events), batch_size):
            batch = events[batch_start:batch_start + batch_size]
            batch_num = batch_start // batch_size + 1
            
            results = classify_batch(batch)
            processed_indices = set()
            
            for res in results:
                idx = res.get("i", 1) - 1
                if 0 <= idx < len(batch):
                    processed_indices.add(idx)
                    event = dict(batch[idx])
                    event["artists"] = res.get("artists") or []
                    event["genre"] = res.get("genre", "Don't Box Me!")
                    event["event_type"] = res.get("event_type", "Other")
                    event["delete"] = res.get("delete", False)
                    event["is_event"] = res.get("is_event", False)
                    event["has_more"] = res.get("has_more", False)
                    event["event_bio"] = res.get("event_bio")
                    classified.append(event)
            
            for idx, e in enumerate(batch):
                if idx not in processed_indices:
                    e2 = dict(e)
                    e2["artists"] = []
                    e2["genre"] = "Don't Box Me!"
                    e2["event_type"] = "Other"
                    e2["delete"] = False
                    e2["is_event"] = True
                    classified.append(e2)
            
            artists_count = sum(len(r.get("artists",[])) for r in results)
            deletes_count = sum(1 for r in results if r.get("delete"))
            events_count = sum(1 for r in results if r.get("is_event"))
            more_count = sum(1 for r in results if r.get("has_more"))
            log.info(f"  [{batch_num}/{total_batches}] ✅ {artists_count} artists, {events_count} events, {deletes_count} delete, {more_count} +more")
        
        total_artist_rows = sum(max(1, len(e.get("artists",[]))) for e in classified if e.get("artists"))
        total_events_only = sum(1 for e in classified if e.get("is_event") and not e.get("delete"))
        total_delete = sum(1 for e in classified if e.get("delete"))
        total_has_more = sum(1 for e in classified if e.get("has_more"))
        log.info(f"\n   Classification done: {total_artist_rows} artist rows, {total_events_only} events, {total_delete} DELETE, {total_has_more} with +more\n")
        
        # ── Step 2.5: Lineup Search for "+more" events ──
        if total_has_more > 0:
            log.info(f"🔎 Phase 1.5: Searching full lineups for {total_has_more} events with '+more'...")
            for e in classified:
                if e.get("has_more") and e.get("artists"):
                    additional = search_event_lineup(
                        e["title"], e["venue"], e["city"], e["artists"]
                    )
                    if additional:
                        for name in additional:
                            name = name.strip()
                            if name and len(name) > 1 and normalize(name) not in {normalize(a) for a in e["artists"]}:
                                e["artists"].append(name)
                        log.info(f"  +{len(additional)} found for: {e['title'][:50]}")
            total_artist_rows = sum(max(1, len(e.get("artists",[]))) for e in classified if e.get("artists"))
            log.info(f"   Updated total: {total_artist_rows} artist rows\n")
        
        # 💾 Save after Phase 1 + 1.5
        cp.save(phase=1, classified=classified)
    
    # ── Step 3: Spotify Verification ──
    # Build unique_artists from classified (needed by later phases regardless)
    unique_artists = {}
    for e in classified:
        for a in e.get("artists", []):
            norm = normalize(a)
            if norm and norm not in unique_artists:
                unique_artists[norm] = {"name": a, "city": e["city"], "genre": e.get("genre", "")}
    
    if _cp_phase >= 2:
        verify_cache = _cp_data["verify_cache"]
        verified = len(verify_cache)
        log.info(f"🔍 Phase 2: ⏩ Restored {verified} verified artists from checkpoint")
    else:
        log.info("🔍 Phase 2: Multi-Platform Verification (Spotify + Deezer + MusicBrainz)...")
        log.info(f"   {len(unique_artists)} unique artists to verify")
        verify_cache = {}
        verified = 0
        
        for i, (norm, info) in enumerate(unique_artists.items()):
            result = verify_artist_multiplatform(info["name"])
            if result:
                verify_cache[norm] = result
                verified += 1
            if (i + 1) % 20 == 0:
                sources_count = {}
                for v in verify_cache.values():
                    for s in v.get("sources", []):
                        sources_count[s] = sources_count.get(s, 0) + 1
                src_str = ", ".join(f"{k}={v}" for k, v in sources_count.items())
                log.info(f"  [{i+1}/{len(unique_artists)}] ✅ {verified} verified ({src_str})")
            time.sleep(0.05)
        
        sources_count = {}
        for v in verify_cache.values():
            for s in v.get("sources", []):
                sources_count[s] = sources_count.get(s, 0) + 1
        src_str = ", ".join(f"{k}={v}" for k, v in sources_count.items())
        log.info(f"   Verified: {verified}/{len(unique_artists)} ({src_str})\n")
        # 💾 Save after Phase 2
        cp.save(phase=2, classified=classified, verify_cache=verify_cache)
    
    
    # ── Step 4: AI Enrichment (bio, locale, URLs, emails — David's 3-pass) ──
    if _cp_phase >= 3:
        enrichment_cache = _cp_data["enrichment_cache"]
        log.info(f"📝 Phase 3: ⏩ Restored {len(enrichment_cache)} enriched artists from checkpoint")
    else:
        log.info("📝 Phase 3: AI Enrichment (David's 3-pass email search)...")
        artists_to_enrich = [info for norm, info in unique_artists.items() if info.get("name")]
        enrichment_cache = {}
        enrich_batch_size = 10
        total_enrichment_batches = (len(artists_to_enrich) + enrich_batch_size - 1) // enrich_batch_size
        
        for batch_start in range(0, len(artists_to_enrich), enrich_batch_size):
            batch = artists_to_enrich[batch_start:batch_start + enrich_batch_size]
            batch_num = batch_start // enrich_batch_size + 1
            
            results = enrich_batch(batch)
            
            for res in results:
                idx = res.get("i", 1) - 1
                if 0 <= idx < len(batch):
                    norm = normalize(batch[idx]["name"])
                    enrichment_cache[norm] = res
            
            log.info(f"  [{batch_num}/{total_enrichment_batches}] ✅ Enriched {len(enrichment_cache)} artists")
        
        log.info(f"   Enrichment done: {len(enrichment_cache)} artists\n")
        # 💾 Save after Phase 3
        cp.save(phase=3, classified=classified, verify_cache=verify_cache, enrichment_cache=enrichment_cache)
    
    # ── Step 3.5: Profile Building (ALL sources) ──
    _skip_35 = False
    if _cp_phase >= 35:
        profile_cache = _cp_data.get("profile_cache", {})
        log.info(f"🌐 Phase 3.5: ⏩ Restored {len(profile_cache)} artist profiles from checkpoint")
        _skip_35 = True

    if not _skip_35:
        log.info("🌐 Phase 3.5: Profile Building (13 sources: MusicBrainz, Last.fm, Wikipedia, Wikidata, iTunes, Discogs, YouTube, Genius, Google KG, Setlist.fm, Serper/DDG, Linktree, Scraping)...")
    profile_cache = {} if not _skip_35 else profile_cache
    mb_links_found = 0
    sites_scraped = 0
    spotify_via_mb = 0
    social_found = 0
    stats = {"lastfm": 0, "wiki": 0, "wikidata": 0, "itunes": 0, "discogs": 0, "youtube": 0, "genius": 0, "ddg": 0, "linktree": 0, "kg": 0, "serper": 0, "setlistfm": 0,
             "bandsintown": 0, "soundcloud": 0, "bandcamp": 0, "deezer_ext": 0, "mb_extras": 0}
    
    # Caches for all sources
    mb_links_cache = {}
    scraped_cache = {}
    lastfm_cache = {}
    wiki_cache = {}
    wikidata_cache = {}
    itunes_cache = {}
    discogs_cache = {}
    youtube_cache = {}
    genius_cache = {}
    kg_cache = {}
    setlistfm_cache = {}
    bandsintown_cache = {}
    soundcloud_cache = {}
    bandcamp_cache = {}
    deezer_ext_cache = {}
    mb_extras_cache = {}
    
    
    # Collect artists that have MBID (from Phase 2 verification)
    artists_with_mbid = []
    if not _skip_35:
        for norm, vf in verify_cache.items():
            if vf.get("mb_id"):
                artists_with_mbid.append((norm, vf))
        log.info(f"   {len(artists_with_mbid)} artists have MBID, {len(unique_artists)} total artists")
    
    # ── Sub-step A: MusicBrainz external links ──
    for i, (norm, vf) in enumerate(artists_with_mbid):
        links = mb_get_links(vf["mb_id"])
        if links:
            mb_links_cache[norm] = links
            mb_links_found += 1
            if "spotify" in links:
                spotify_via_mb += 1
            social_found += sum(1 for k in links if k in ("instagram","facebook","twitter","tiktok","youtube"))
            # Extract extra platform links (RA, AllMusic, Songkick, etc.)
            extras = extract_mb_extra_links(links)
            if extras:
                mb_extras_cache[norm] = extras
                stats["mb_extras"] += 1
        if (i + 1) % 10 == 0:
            log.info(f"  [{i+1}/{len(artists_with_mbid)}] MB links: {mb_links_found} found, Spotify: {spotify_via_mb}")
    
    log.info(f"   ✅ MusicBrainz links: {mb_links_found} artists, {spotify_via_mb} Spotify, {social_found} social, {stats['mb_extras']} extra platforms (RA/AllMusic/etc)")
    
    # ── Sub-step B: All other APIs (parallel per artist) ──
    def _enrich_artist_sources(norm_name_tuple):
        """Fetch data from all APIs for one artist (runs in thread)."""
        norm, artist_name = norm_name_tuple
        result = {"norm": norm}
        
        # Last.fm
        try:
            lfm = lastfm_search(artist_name)
            if lfm:
                result["lastfm"] = lfm
        except Exception as e:
            log.debug(f"Last.fm failed for '{artist_name}': {e}")
        
        # Wikipedia
        try:
            wiki = wikipedia_search(artist_name)
            if wiki:
                result["wiki"] = wiki
        except Exception as e:
            log.debug(f"Wikipedia failed for '{artist_name}': {e}")
        
        # Wikidata (structured bio data)
        try:
            wd = wikidata_search(artist_name)
            if wd:
                result["wikidata"] = wd
        except Exception as e:
            log.debug(f"Wikidata failed for '{artist_name}': {e}")
        
        # iTunes/Apple Music (free, no key)
        try:
            it = itunes_search(artist_name)
            if it:
                result["itunes"] = it
        except Exception as e:
            log.debug(f"iTunes failed for '{artist_name}': {e}")
        
        # Discogs
        try:
            dc = discogs_search(artist_name)
            if dc:
                result["discogs"] = dc
        except Exception as e:
            log.debug(f"Discogs failed for '{artist_name}': {e}")
        
        # YouTube
        try:
            yt = youtube_search(artist_name)
            if yt:
                result["youtube"] = yt
        except Exception as e:
            log.debug(f"YouTube failed for '{artist_name}': {e}")
        
        # Genius
        try:
            gn = genius_search(artist_name)
            if gn:
                result["genius"] = gn
        except Exception as e:
            log.debug(f"Genius failed for '{artist_name}': {e}")
        
        # Google Knowledge Graph
        try:
            kg = google_kg_search(artist_name)
            if kg:
                result["kg"] = kg
        except Exception as e:
            log.debug(f"KG failed for '{artist_name}': {e}")
        
        # Setlist.fm
        try:
            sl = setlistfm_search(artist_name)
            if sl:
                result["setlistfm"] = sl
        except Exception as e:
            log.debug(f"Setlist.fm failed for '{artist_name}': {e}")
        
        # Bandsintown (events/shows)
        try:
            bit = bandsintown_search(artist_name)
            if bit:
                result["bandsintown"] = bit
        except Exception as e:
            log.debug(f"Bandsintown failed for '{artist_name}': {e}")
        
        # SoundCloud
        try:
            sc = soundcloud_search(artist_name)
            if sc:
                result["soundcloud"] = sc
        except Exception as e:
            log.debug(f"SoundCloud failed for '{artist_name}': {e}")
        
        # Bandcamp
        try:
            bc = bandcamp_search(artist_name)
            if bc:
                result["bandcamp"] = bc
        except Exception as e:
            log.debug(f"Bandcamp failed for '{artist_name}': {e}")
        
        return result
    
    # Build list of (norm, name) pairs for all unique artists
    artist_pairs = [(norm, unique_artists[norm]["name"]) for norm in unique_artists] if not _skip_35 else []
    api_results = []
    
    if not _skip_35:
        log.info(f"   Fetching from Last.fm, Wikipedia, Wikidata, iTunes, Discogs, YouTube, Genius, KG, Setlist.fm, Bandsintown, SoundCloud, Bandcamp for {len(artist_pairs)} artists...")
        with ThreadPoolExecutor(max_workers=4) as pool:
            api_results = list(pool.map(_enrich_artist_sources, artist_pairs))
    
    # Store results in caches
    for res in api_results:
        norm = res["norm"]
        if "lastfm" in res:
            lastfm_cache[norm] = res["lastfm"]
            stats["lastfm"] += 1
        if "wiki" in res:
            wiki_cache[norm] = res["wiki"]
            stats["wiki"] += 1
        if "wikidata" in res:
            wikidata_cache[norm] = res["wikidata"]
            stats["wikidata"] += 1
        if "itunes" in res:
            itunes_cache[norm] = res["itunes"]
            stats["itunes"] += 1
        if "discogs" in res:
            discogs_cache[norm] = res["discogs"]
            stats["discogs"] += 1
        if "youtube" in res:
            youtube_cache[norm] = res["youtube"]
            stats["youtube"] += 1
        if "genius" in res:
            genius_cache[norm] = res["genius"]
            stats["genius"] += 1
        if "kg" in res:
            kg_cache[norm] = res["kg"]
            stats["kg"] += 1
        if "setlistfm" in res:
            setlistfm_cache[norm] = res["setlistfm"]
            stats["setlistfm"] += 1
        if "bandsintown" in res:
            bandsintown_cache[norm] = res["bandsintown"]
            stats["bandsintown"] += 1
        if "soundcloud" in res:
            soundcloud_cache[norm] = res["soundcloud"]
            stats["soundcloud"] += 1
        if "bandcamp" in res:
            bandcamp_cache[norm] = res["bandcamp"]
            stats["bandcamp"] += 1
    
    if not _skip_35:
        log.info(f"   ✅ APIs done: Last.fm={stats['lastfm']}, Wiki={stats['wiki']}, Wikidata={stats['wikidata']}, iTunes={stats['itunes']}, Discogs={stats['discogs']}, YouTube={stats['youtube']}, Genius={stats['genius']}, KG={stats['kg']}, Setlist.fm={stats['setlistfm']}")
        log.info(f"   ✅ NEW APIs: Bandsintown={stats['bandsintown']}, SoundCloud={stats['soundcloud']}, Bandcamp={stats['bandcamp']}")
    
    # ── Sub-step C: Scrape official websites + discover via DuckDuckGo ──
    websites_to_scrape = []
    for norm, links in mb_links_cache.items():
        if links.get("website"):
            websites_to_scrape.append((norm, links["website"]))
    # Also check Wikidata websites
    for norm, wd in wikidata_cache.items():
        if norm not in [w[0] for w in websites_to_scrape] and wd.get("websites"):
            websites_to_scrape.append((norm, wd["websites"][0]))
    # Also check Discogs URLs
    for norm, dc in discogs_cache.items():
        if norm not in [w[0] for w in websites_to_scrape] and dc.get("urls"):
            for durl in dc["urls"]:
                if durl and "discogs.com" not in durl and "facebook.com" not in durl:
                    cat = classify_url(durl)
                    if cat == "website":
                        websites_to_scrape.append((norm, durl))
                        break
    # Also check enrichment URLs for official sites
    for norm, enrich in enrichment_cache.items():
        if norm not in [w[0] for w in websites_to_scrape]:
            for uk in ["url1", "url2", "url3"]:
                uu = enrich.get(uk)
                if uu and "deezer.com" not in uu and "spotify.com" not in uu and "musicbrainz.org" not in uu:
                    cat = classify_url(uu)
                    if cat == "website":
                        websites_to_scrape.append((norm, uu))
                        break
    
    log.info(f"   {len(websites_to_scrape)} official websites to scrape")
    for i, (norm, site_url) in enumerate(websites_to_scrape):
        scraped = scrape_official_site(site_url)
        if scraped:
            scraped_cache[norm] = scraped
            sites_scraped += 1
            # Check for Linktree links
            for platform, pu in scraped.items():
                if isinstance(pu, str) and "linktr.ee" in pu:
                    lt_data = scrape_linktree(pu)
                    if lt_data:
                        for k2, v2 in lt_data.items():
                            if k2 not in scraped:
                                scraped[k2] = v2
                        stats["linktree"] += 1
        if (i + 1) % 10 == 0:
            log.info(f"  [{i+1}/{len(websites_to_scrape)}] Scraped: {sites_scraped} sites")
    
    # Serper.dev (Google Search): discover websites for artists without one — much better than DDG
    artists_without_site = [norm for norm in unique_artists if norm not in scraped_cache and norm not in [w[0] for w in websites_to_scrape]] if not _skip_35 else []
    # Also add websites from KG
    for norm, kg in kg_cache.items():
        if kg.get("website") and norm not in [w[0] for w in websites_to_scrape]:
            ws_url = kg["website"]
            scraped = scrape_official_site(ws_url)
            if scraped:
                scraped_cache[norm] = scraped
                sites_scraped += 1
    if artists_without_site:
        search_fn = serper_search if SERPER_KEY else ddg_search
        search_name = "Serper (Google)" if SERPER_KEY else "DuckDuckGo"
        log.info(f"   {search_name}: searching for {len(artists_without_site)} artists without website...")
        serper_found = 0
        for norm in artists_without_site[:50]:  # Serper has higher limits
            name = unique_artists[norm]["name"]
            if SERPER_KEY:
                results = serper_search(f"{name} music artist official website")
                found_urls = [r["link"] for r in results if r.get("link")]
            else:
                found_urls = ddg_search(f"{name} music artist official website")
            for durl in found_urls[:5]:
                if not durl:
                    continue
                cat = classify_url(durl)
                if cat == "website":
                    scraped = scrape_official_site(durl)
                    if scraped:
                        scraped_cache[norm] = scraped
                        sites_scraped += 1
                        serper_found += 1
                    break
                elif "linktr.ee" in durl:
                    lt_data = scrape_linktree(durl)
                    if lt_data:
                        scraped_cache[norm] = lt_data
                        stats["linktree"] += 1
                        serper_found += 1
                    break
        stats["serper" if SERPER_KEY else "ddg"] = serper_found
        log.info(f"   ✅ {search_name} found: {serper_found} new sites/linktrees")
    
    if not _skip_35:
        log.info(f"   ✅ Websites scraped: {sites_scraped}, Linktrees: {stats['linktree']}")
    
    # ── Sub-step D: Build merged profiles ──
    for norm in (unique_artists if not _skip_35 else {}):
        vf = verify_cache.get(norm)
        enrich = enrichment_cache.get(norm, {})
        mb_links = mb_links_cache.get(norm)
        scraped = scraped_cache.get(norm)
        
        profile = build_profile(unique_artists[norm]["name"], vf, enrich, mb_links, scraped)
        
        # Merge Last.fm data
        lfm = lastfm_cache.get(norm)
        if lfm:
            if lfm.get("lastfm_url") and "lastfm" not in profile["platforms"]:
                profile["platforms"]["lastfm"] = lfm["lastfm_url"]
                if lfm["lastfm_url"] not in profile["urls"]:
                    profile["urls"].append(lfm["lastfm_url"])
            if lfm.get("bio") and (not profile.get("bio") or len(profile["bio"]) < len(lfm["bio"])):
                profile["bio"] = lfm["bio"]
            if lfm.get("listeners"):
                profile["listeners"] = lfm["listeners"]
            if lfm.get("tags"):
                profile["lastfm_tags"] = lfm["tags"]
        
        # Merge Wikipedia data
        wiki = wiki_cache.get(norm)
        if wiki:
            if wiki.get("wiki_url") and "wikipedia" not in profile["platforms"]:
                profile["platforms"]["wikipedia"] = wiki["wiki_url"]
                if wiki["wiki_url"] not in profile["urls"]:
                    profile["urls"].append(wiki["wiki_url"])
            if wiki.get("bio") and (not profile.get("bio") or len(profile["bio"]) < len(wiki["bio"])):
                profile["bio"] = wiki["bio"]
            if wiki.get("thumbnail"):
                profile["photo"] = wiki["thumbnail"]
        
        # Merge Wikidata (structured fields)
        wd = wikidata_cache.get(norm)
        if wd:
            if wd.get("born"):
                profile["born"] = wd["born"]
            if wd.get("birthplace"):
                profile["birthplace"] = wd["birthplace"]
            if wd.get("years_active"):
                profile["years_active"] = wd["years_active"]
            if wd.get("instruments"):
                profile["instruments"] = wd["instruments"]
            if wd.get("labels"):
                profile["record_labels"] = wd["labels"]
            if wd.get("occupations"):
                profile["occupations"] = wd["occupations"]
            if wd.get("genres"):
                profile["wikidata_genres"] = wd["genres"]
            if wd.get("websites"):
                for wsite in wd["websites"]:
                    if "website" not in profile["platforms"]:
                        profile["platforms"]["website"] = wsite
                        if wsite not in profile["urls"]:
                            profile["urls"].append(wsite)
        
        # Merge iTunes
        it = itunes_cache.get(norm)
        if it:
            if it.get("itunes_url") and "apple_music" not in profile["platforms"]:
                profile["platforms"]["apple_music"] = it["itunes_url"]
                if it["itunes_url"] not in profile["urls"]:
                    profile["urls"].append(it["itunes_url"])
            if it.get("genre"):
                profile["itunes_genre"] = it["genre"]
        
        # Merge Discogs
        dc = discogs_cache.get(norm)
        if dc:
            if dc.get("discogs_url") and "discogs" not in profile["platforms"]:
                profile["platforms"]["discogs"] = dc["discogs_url"]
                if dc["discogs_url"] not in profile["urls"]:
                    profile["urls"].append(dc["discogs_url"])
            if dc.get("real_name"):
                profile["real_name"] = dc["real_name"]
            # Also scrape Discogs URLs for more links
            if dc.get("urls"):
                for durl in dc["urls"]:
                    if durl and durl not in profile["urls"]:
                        cat = classify_url(durl)
                        if cat not in profile["platforms"]:
                            profile["platforms"][cat] = durl
                            profile["urls"].append(durl)
        
        # Merge YouTube
        yt = youtube_cache.get(norm)
        if yt:
            if yt.get("channel_url") and "youtube" not in profile["platforms"]:
                profile["platforms"]["youtube"] = yt["channel_url"]
                if yt["channel_url"] not in profile["urls"]:
                    profile["urls"].append(yt["channel_url"])
        
        # Merge Genius
        gn = genius_cache.get(norm)
        if gn:
            if gn.get("genius_url") and "genius" not in profile["platforms"]:
                profile["platforms"]["genius"] = gn["genius_url"]
                if gn["genius_url"] not in profile["urls"]:
                    profile["urls"].append(gn["genius_url"])
            if gn.get("top_songs"):
                profile["top_songs"] = gn["top_songs"]
            if gn.get("genius_image") and not profile.get("photo"):
                profile["photo"] = gn["genius_image"]
        
        # Add photo from verification (Deezer)
        if vf and vf.get("deezer_picture") and not profile.get("photo"):
            profile["photo"] = vf["deezer_picture"]
        
        # Merge Google Knowledge Graph
        kg = kg_cache.get(norm)
        if kg:
            if kg.get("description"):
                profile["kg_description"] = kg["description"]
            if kg.get("types"):
                profile["kg_types"] = kg["types"]
            if kg.get("is_musician"):
                profile["kg_confirmed_musician"] = True
            if kg.get("bio") and (not profile.get("bio") or len(profile["bio"]) < len(kg["bio"])):
                profile["bio"] = kg["bio"]
            if kg.get("image") and not profile.get("photo"):
                profile["photo"] = kg["image"]
            if kg.get("website") and "website" not in profile["platforms"]:
                profile["platforms"]["website"] = kg["website"]
                if kg["website"] not in profile["urls"]:
                    profile["urls"].append(kg["website"])
        
        # Merge Setlist.fm
        sl = setlistfm_cache.get(norm)
        if sl:
            profile["setlistfm_url"] = sl.get("url", "")
            profile["total_setlists"] = sl.get("total_setlists", 0)
            profile["is_active_performer"] = sl.get("is_active", False)
            if sl.get("disambiguation"):
                profile["disambiguation"] = sl["disambiguation"]
            if sl.get("url") and sl["url"] not in profile["urls"]:
                profile["urls"].append(sl["url"])
        
        # Merge Bandsintown (events/touring)
        bit = bandsintown_cache.get(norm)
        if bit:
            if bit.get("bandsintown_url") and "bandsintown" not in profile["platforms"]:
                profile["platforms"]["bandsintown"] = bit["bandsintown_url"]
                if bit["bandsintown_url"] not in profile["urls"]:
                    profile["urls"].append(bit["bandsintown_url"])
            if bit.get("tracker_count"):
                profile["bandsintown_trackers"] = bit["tracker_count"]
            if bit.get("upcoming_events"):
                profile["upcoming_events"] = bit["upcoming_events"]
            if bit.get("is_touring"):
                profile["is_touring"] = True
            if bit.get("events"):
                profile["upcoming_shows"] = bit["events"]
            if bit.get("facebook_url") and "facebook" not in profile["platforms"]:
                profile["platforms"]["facebook"] = bit["facebook_url"]
        
        # Merge SoundCloud
        sc = soundcloud_cache.get(norm)
        if sc:
            if sc.get("soundcloud_url") and "soundcloud" not in profile["platforms"]:
                profile["platforms"]["soundcloud"] = sc["soundcloud_url"]
                if sc["soundcloud_url"] not in profile["urls"]:
                    profile["urls"].append(sc["soundcloud_url"])
            if sc.get("followers"):
                profile["soundcloud_followers"] = sc["followers"]
            if sc.get("track_count"):
                profile["soundcloud_tracks"] = sc["track_count"]
            if sc.get("description") and (not profile.get("bio") or len(profile["bio"]) < 50):
                profile["bio"] = sc["description"]
            if sc.get("verified"):
                profile["soundcloud_verified"] = True
        
        # Merge Bandcamp
        bc = bandcamp_cache.get(norm)
        if bc:
            if bc.get("bandcamp_url") and "bandcamp" not in profile["platforms"]:
                profile["platforms"]["bandcamp"] = bc["bandcamp_url"]
                if bc["bandcamp_url"] not in profile["urls"]:
                    profile["urls"].append(bc["bandcamp_url"])
            if bc.get("location") and not profile.get("birthplace"):
                profile["location"] = bc["location"]
            if bc.get("genre"):
                profile["bandcamp_genre"] = bc["genre"]
        
        # Merge Deezer Extended data (albums, top tracks, related artists)
        vf_deezer = vf.get("deezer_url", "") if vf else ""
        if vf_deezer and norm not in deezer_ext_cache:
            try:
                dext = deezer_extended_search(vf_deezer)
                if dext:
                    deezer_ext_cache[norm] = dext
                    stats["deezer_ext"] += 1
            except: pass
        dext = deezer_ext_cache.get(norm)
        if dext:
            if dext.get("top_tracks") and not profile.get("top_songs"):
                profile["top_songs"] = [t["title"] for t in dext["top_tracks"]]
            if dext.get("albums"):
                profile["discography"] = dext["albums"]
                profile["album_count"] = dext.get("album_count", len(dext["albums"]))
            if dext.get("related_artists"):
                profile["related_artists"] = dext["related_artists"]
        
        # Merge MB Extra Links (RA, AllMusic, Songkick, RYM, etc.)
        mb_ext = mb_extras_cache.get(norm)
        if mb_ext:
            for platform, url in mb_ext.items():
                if platform not in profile["platforms"]:
                    profile["platforms"][platform] = url
                    if url not in profile["urls"]:
                        profile["urls"].append(url)
        
        # ── Genre Resolution ──
        # Consolidate genre from all sources into a single field
        genre_candidates = []  # [(genre_string, source, priority)]
        
        # Priority 1: Spotify genres (most reliable, curated by Spotify)
        if vf and vf.get("spotify_genres"):
            for sg in vf["spotify_genres"]:
                mapped = validate_genre(sg)
                if mapped != "Don't Box Me!":
                    genre_candidates.append((mapped, "spotify", 1))
                    break
        
        # Priority 2: iTunes genre
        if profile.get("itunes_genre"):
            mapped = validate_genre(profile["itunes_genre"])
            if mapped != "Don't Box Me!":
                genre_candidates.append((mapped, "itunes", 2))
        
        # Priority 3: AI classification (context-aware from event title)
        ai_genre = unique_artists[norm].get("genre", "")
        if ai_genre:
            mapped = validate_genre(ai_genre)
            if mapped != "Don't Box Me!":
                genre_candidates.append((mapped, "ai_classify", 3))
        
        # Priority 4: Last.fm tags
        if profile.get("lastfm_tags"):
            for tag in profile["lastfm_tags"]:
                mapped = validate_genre(tag)
                if mapped != "Don't Box Me!":
                    genre_candidates.append((mapped, "lastfm", 4))
                    break
        
        # Priority 5: MusicBrainz tags
        if vf and vf.get("mb_tags"):
            for tag in vf["mb_tags"]:
                mapped = validate_genre(tag)
                if mapped != "Don't Box Me!":
                    genre_candidates.append((mapped, "musicbrainz", 5))
                    break
        
        # Priority 6: Wikidata genres
        if profile.get("wikidata_genres"):
            for wg in profile["wikidata_genres"]:
                mapped = validate_genre(wg)
                if mapped != "Don't Box Me!":
                    genre_candidates.append((mapped, "wikidata", 6))
                    break
        
        # Priority 7: Bandcamp genre
        if profile.get("bandcamp_genre"):
            mapped = validate_genre(profile["bandcamp_genre"])
            if mapped != "Don't Box Me!":
                genre_candidates.append((mapped, "bandcamp", 7))
        
        # Pick the best genre (lowest priority number = highest reliability)
        if genre_candidates:
            genre_candidates.sort(key=lambda x: x[2])
            profile["genre"] = genre_candidates[0][0]
            profile["genre_source"] = genre_candidates[0][1]
            # Also store all unique genres found across sources
            seen = set()
            profile["all_genres"] = []
            for g, src, _ in genre_candidates:
                if g not in seen:
                    seen.add(g)
                    profile["all_genres"].append(g)
        else:
            profile["genre"] = "Don't Box Me!"
            profile["genre_source"] = None
            profile["all_genres"] = []
        
        profile_cache[norm] = profile
    
    if not _skip_35:
        log.info(f"   ✅ Profiles built: {len(profile_cache)}")
        log.info(f"   📊 Summary: MB_links={mb_links_found}, Spotify(MB)={spotify_via_mb}, Scraped={sites_scraped}")
        log.info(f"   📊 APIs: Last.fm={stats['lastfm']}, Wiki={stats['wiki']}, Wikidata={stats['wikidata']}, iTunes={stats['itunes']}")
        log.info(f"   📊 APIs: Discogs={stats['discogs']}, YouTube={stats['youtube']}, Genius={stats['genius']}, KG={stats['kg']}, Setlist.fm={stats['setlistfm']}")
        log.info(f"   📊 NEW:  Bandsintown={stats['bandsintown']}, SoundCloud={stats['soundcloud']}, Bandcamp={stats['bandcamp']}, Deezer_ext={stats['deezer_ext']}, MB_extras={stats['mb_extras']}")
        log.info(f"   📊 Discovery: Serper/DDG={stats.get('serper',0)+stats.get('ddg',0)}, Linktree={stats['linktree']}\n")
    if not _skip_35:
        # 💾 Save after Phase 3.5 (profile building is the most expensive step)
        # Flatten sub-caches into checkpoint — only the merged profile_cache is needed for Phase 4+5
        cp.save(phase=35, classified=classified, verify_cache=verify_cache,
            enrichment_cache=enrichment_cache, profile_cache=profile_cache)
    
    # ── Step 4: URL Verification + Email Validation ──
    if _cp_phase >= 35:
        profile_cache = _cp_data.get("profile_cache", profile_cache)
    log.info("🔗 Phase 4: URL Verification + Email Validation...")
    urls_checked = 0
    urls_valid = 0
    urls_removed = 0
    emails_checked = 0
    emails_valid = 0
    emails_removed = 0
    
    # Known-good domains (already verified by APIs)
    _TRUSTED_DOMAINS = {"deezer.com", "spotify.com", "musicbrainz.org", "apple.com", "music.apple.com",
                         "soundcloud.com", "bandcamp.com", "youtube.com", "youtu.be",
                         "instagram.com", "facebook.com", "twitter.com", "x.com", "tiktok.com",
                         "discogs.com", "allmusic.com", "songkick.com", "last.fm", "wikidata.org", "wikipedia.org",
                         "genius.com", "setlist.fm", "audiomack.com"}
    
    def _is_trusted(url):
        """Check if URL is from a trusted domain (skip HTTP check)."""
        try:
            host = urllib.parse.urlparse(url).hostname or ""
            return any(host.endswith(d) for d in _TRUSTED_DOMAINS)
        except:
            return False
    
    # Validate all profile URLs
    for norm, profile in profile_cache.items():
        valid_platforms = {}
        for platform, url in profile.get("platforms", {}).items():
            if url:
                urls_checked += 1
                if _is_trusted(url):
                    urls_valid += 1
                    valid_platforms[platform] = url
                elif verify_url(url):
                    urls_valid += 1
                    valid_platforms[platform] = url
                else:
                    urls_removed += 1
        profile["platforms"] = valid_platforms
        # Rebuild flat URL list
        profile["urls"] = list(valid_platforms.values())
        
        # Validate profile emails
        valid_emails = []
        for em in profile.get("emails", []):
            emails_checked += 1
            if validate_email(em):
                emails_valid += 1
                valid_emails.append(em)
            else:
                emails_removed += 1
        profile["emails"] = valid_emails
    
    log.info(f"   URLs: {urls_valid}/{urls_checked} valid ({urls_removed} removed)")
    log.info(f"   Emails: {emails_valid}/{emails_checked} valid ({emails_removed} removed)\n")
    
    # ── Step 5: Write Output (handle multi-artist row splitting) ──
    log.info("📝 Writing output Excel...")
    
    delete_fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
    event_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # green
    med_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # yellow
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # red
    
    output_rows = []
    
    for e in classified:
        artists = e.get("artists", [])
        
        if e.get("delete"):
            output_rows.append({
                "id": "DELETE", "city": e["city"], "venue": e["venue"], "title": e["title"],
                "artist": None, "genre": e.get("genre"), "event_type": sanitize_event_type(e.get("event_type")),
                "delete": True, "is_event": True, "event_bio": e.get("event_bio")
            })
        elif not artists or e.get("is_event"):
            output_rows.append({
                "id": None, "city": e["city"], "venue": e["venue"], "title": e["title"],
                "artist": "Event", "genre": e.get("genre"), "event_type": sanitize_event_type(e.get("event_type")),
                "delete": False, "is_event": True, "event_bio": e.get("event_bio")
            })
        elif len(artists) == 1:
            output_rows.append({
                "id": None, "city": e["city"], "venue": e["venue"], "title": e["title"],
                "artist": artists[0], "genre": e.get("genre"), "event_type": sanitize_event_type(e.get("event_type")),
                "delete": False, "is_event": False
            })
        else:
            for artist_name in artists:
                output_rows.append({
                    "id": None, "city": e["city"], "venue": e["venue"], "title": e["title"],
                    "artist": artist_name, "genre": e.get("genre"), "event_type": sanitize_event_type(e.get("event_type")),
                    "delete": False, "is_event": False
                })
    
    data_start_row = 4
    confidence_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}
    
    for out_idx, out in enumerate(output_rows):
        r = data_start_row + out_idx
        
        if out.get("delete"):
            ws.cell(r, 1).value = "DELETE"
            ws.cell(r, 1).fill = delete_fill
        
        ws.cell(r, 2).value = out["city"]
        ws.cell(r, 3).value = out["venue"]
        ws.cell(r, 4).value = out["title"]
        
        artist_name = out.get("artist")
        is_event = out.get("is_event", False)
        
        if artist_name and artist_name != "Event" and not is_event:
            norm = normalize(artist_name)
            vf = verify_cache.get(norm)
            profile = profile_cache.get(norm, {})
            enrich = enrichment_cache.get(norm, {})
            
            # Artist name — use profile/verified name
            ws.cell(r, 6).value = profile.get("name") or (vf["name"] if vf else artist_name)
            
            # Genre — priority: Spotify > AI > MusicBrainz
            genre = best_genre(out.get("genre", "Don't Box Me!"), vf)
            ws.cell(r, 7).value = genre
            
            # Event Type
            ws.cell(r, 8).value = out.get("event_type", "Live Music Performance")
            
            # Bio — from profile (merged from enrichment + scraped site)
            bio = profile.get("bio") or enrich.get("bio")
            if bio: ws.cell(r, 9).value = bio
            
            # Locale — from profile (merged from enrichment + MusicBrainz)
            locale = profile.get("locale", {})
            if locale.get("city"): ws.cell(r, 10).value = locale["city"]
            if locale.get("state"): ws.cell(r, 11).value = locale["state"]
            if locale.get("country"): ws.cell(r, 12).value = locale["country"]
            
            # URLs — from profile (merged + deduped + validated platforms)
            platforms = profile.get("platforms", {})
            # Priority order: Spotify, Deezer, Apple Music, SoundCloud, YouTube, website, social
            url_priority = ["spotify", "deezer", "apple_music", "soundcloud", "youtube", "bandcamp",
                          "website", "instagram", "facebook", "twitter", "tiktok"]
            ordered_urls = []
            for p in url_priority:
                if p in platforms:
                    ordered_urls.append(platforms[p])
            # Add any remaining platforms not in priority list
            for p, u in platforms.items():
                if u not in ordered_urls:
                    ordered_urls.append(u)
            for ui, url_val in enumerate(ordered_urls[:3]):
                ws.cell(r, 13 + ui).value = url_val
            
            # Emails — from profile (merged + validated), with label annotation
            for ei, em in enumerate(profile.get("emails", [])[:3]):
                label = email_label(em)
                cell_val = f"{em} [{label}]" if label and label != "general" else em
                ws.cell(r, 16 + ei).value = cell_val
            
            # Confidence Score (column 19) — now considers profile richness
            conf = calc_confidence(vf, enrich)
            # Boost to HIGH if profile has many platforms or confirmed by multiple sources
            if conf == "MEDIUM":
                if len(platforms) >= 4:
                    conf = "HIGH"
                elif norm in wiki_cache or norm in lastfm_cache:
                    conf = "HIGH"  # Wikipedia/Last.fm = well-known artist
                elif norm in kg_cache and kg_cache[norm].get("is_musician"):
                    conf = "HIGH"  # Google KG confirmed as musician
                elif norm in setlistfm_cache and setlistfm_cache[norm].get("total_setlists", 0) > 5:
                    conf = "HIGH"  # Active performer with 5+ setlists
                elif profile.get("is_active"):
                    conf = "HIGH"  # Has upcoming events from official site
            confidence_counts[conf] = confidence_counts.get(conf, 0) + 1
            ws.cell(r, 19).value = conf
            if conf == "HIGH":
                ws.cell(r, 19).fill = high_fill
            elif conf == "MEDIUM":
                ws.cell(r, 19).fill = med_fill
            else:
                ws.cell(r, 19).fill = low_fill
        else:
            ws.cell(r, 6).value = "Event" if is_event else artist_name
            ws.cell(r, 7).value = validate_genre(out.get("genre", "Don't Box Me!"))
            ws.cell(r, 8).value = out.get("event_type", "Other")
            # Event bio — one-sentence description of the event
            event_bio = out.get("event_bio")
            if event_bio and is_event:
                ws.cell(r, 9).value = event_bio
            if is_event and not out.get("delete"):
                ws.cell(r, 6).fill = event_fill
    
    # Add header for Confidence column
    ws.cell(3, 19).value = "Confidence"
    ws.cell(3, 19).font = Font(bold=True)
    
    # Dump rich profiles to JSON
    rich_profiles = []
    for norm_name, profile in profile_cache.items():
        conf = "LOW"
        if norm_name in verify_cache:
            enrich = enrichment_cache.get(norm_name, {})
            conf = calc_confidence(verify_cache.get(norm_name), enrich)
            if conf == "MEDIUM":
                if len(profile.get("platforms", {})) >= 4 or norm_name in wiki_cache or norm_name in lastfm_cache:
                    conf = "HIGH"
                elif norm_name in kg_cache and kg_cache[norm_name].get("is_musician"): conf = "HIGH"
                elif norm_name in setlistfm_cache and setlistfm_cache[norm_name].get("total_setlists", 0) > 5: conf = "HIGH"
        profile["confidence"] = conf
        
        # Item 5: Rank emails by label priority (validator.py prefix analysis)
        profile["emails"] = rank_emails(
            profile.get("emails", []),
        )
        
        # Item 6: Profile score and tier
        score, tier = calc_profile_score(profile)
        profile["profile_score"] = score
        profile["profile_tier"] = tier
        
        rich_profiles.append(profile)
    
    if dry_run:
        log.info("\n🔍 DRY-RUN mode: skipping file writes and Supabase sync")
    else:
        with open("profiles_rich.json", "w", encoding="utf-8") as f:
            json.dump(rich_profiles, f, indent=2, ensure_ascii=False)
        
        wb.save(OUTPUT_FILE)
        
        # Item 7: Supabase upload via supabase_uploader.py
        upload_result = upload_profiles(rich_profiles, dry_run=dry_run)
        log.info(f"   📤 Supabase: {upload_result['uploaded']}/{upload_result['total']} uploaded")
        
        # Item 9: Output validation summary report
        validate_output_report(rich_profiles)
    
    # Clean up checkpoint — we're done!
    cp.delete()
    
    log.info(f"\n{'='*60}")
    log.info(f"🎉 Done!{' (DRY RUN — no files written)' if dry_run else ''}")
    log.info(f"   Events processed: {len(classified)}")
    log.info(f"   Output rows: {len(output_rows)} (including multi-artist splits)")
    log.info(f"   Artists: {total_artist_rows}")
    log.info(f"   Events (no artist): {total_events_only}")
    log.info(f"   Verified: {verified} ({src_str})")
    log.info(f"   MB Links: {mb_links_found}, Spotify(MB)={spotify_via_mb}, Social={social_found}")
    log.info(f"   APIs: Last.fm={stats['lastfm']}, Wiki={stats['wiki']}, Wikidata={stats['wikidata']}, iTunes={stats['itunes']}")
    log.info(f"   APIs: Discogs={stats['discogs']}, YouTube={stats['youtube']}, Genius={stats['genius']}")
    log.info(f"   Scraped: {sites_scraped} websites, {stats.get('linktree',0)} linktrees, DDG={stats.get('ddg',0)}")
    log.info(f"   URLs: {urls_valid} valid / {urls_removed} removed")
    log.info(f"   Emails: {emails_valid} valid / {emails_removed} removed")
    log.info(f"   Confidence: 🟢 {confidence_counts.get('HIGH',0)} HIGH, 🟡 {confidence_counts.get('MEDIUM',0)} MEDIUM, 🔴 {confidence_counts.get('LOW',0)} LOW")
    log.info(f"   Marked DELETE: {total_delete}")
    if not dry_run:
        log.info(f"   Output saved: {OUTPUT_FILE}")
    log.info(f"{'='*60}")

if __name__ == "__main__":
    main()
