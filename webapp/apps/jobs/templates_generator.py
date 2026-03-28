"""
Generate downloadable Excel template files for each source type.
Templates are generated on-the-fly using openpyxl.
"""
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── Shared styling ──────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
HEADER_FILL = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
EXAMPLE_FONT = Font(name="Calibri", italic=True, color="888888", size=11)
NOTE_FONT = Font(name="Calibri", color="6C5CE7", size=10)
THIN_BORDER = Border(
    bottom=Side(style="thin", color="DDDDDD"),
)


def _style_header(ws, row, cols):
    """Apply header styling to a row."""
    for col in range(1, cols + 1):
        cell = ws.cell(row, col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def _auto_width(ws, min_width=15):
    """Auto-fit column widths."""
    for col in ws.columns:
        # Skip merged cells that don't have column_letter
        first_cell = col[0]
        if not hasattr(first_cell, 'column_letter'):
            continue
        max_len = max((len(str(c.value or "")) for c in col if hasattr(c, 'column_letter')), default=0)
        ws.column_dimensions[first_cell.column_letter].width = max(max_len + 4, min_width)


# ── Excel Source Template (Events) ──────────────────────────────────
def generate_events_template() -> BytesIO:
    """
    Template for ExcelSource — event-based data.
    Sheet: "base info", data starts row 4, columns B/C/D.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "base info"

    # Title rows (1-3)
    ws.merge_cells("A1:D1")
    ws["A1"] = "Groovon — Events Template"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="6C5CE7")

    ws.merge_cells("A2:D2")
    ws["A2"] = "Fill in your event data starting from Row 4. Do NOT modify the header row (Row 3)."
    ws["A2"].font = NOTE_FONT

    # Header row (row 3)
    headers = {
        "A": "#",
        "B": "City",
        "C": "Venue",
        "D": "Title / Event Name",
    }
    for col_letter, label in headers.items():
        cell = ws[f"{col_letter}3"]
        cell.value = label
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Example rows (row 4-6)
    examples = [
        (1, "Berlin", "Lido", "Indie Night — Mashrou' Leila + Cairokee"),
        (2, "Cairo", "Cairo Jazz Club", "Jazz Fusion Night"),
        (3, "Dubai", "Hard Rock Cafe", "Wegz Live in Dubai"),
    ]
    for i, (num, city, venue, title) in enumerate(examples, start=4):
        ws.cell(i, 1, num).font = EXAMPLE_FONT
        ws.cell(i, 2, city).font = EXAMPLE_FONT
        ws.cell(i, 3, venue).font = EXAMPLE_FONT
        ws.cell(i, 4, title).font = EXAMPLE_FONT

    _auto_width(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Artist List Source Template ──────────────────────────────────────
def generate_artist_list_template() -> BytesIO:
    """
    Template for ArtistListSource — pre-extracted artist data.
    Sheet: "Collection cleaned" or auto-detect by "Artist" header.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Collection cleaned"

    # Headers (row 1)
    headers = [
        "Artist",
        "Genre",
        "Bio",
        "Email 1",
        "Email 2",
        "Email 3",
        "Locale City",
        "Local State",
        "Local Country",
        "URL 1",
        "URL 2",
        "URL 3",
        "List",   # performing city
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(1, i, h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Example rows
    examples = [
        [
            "Cairokee", "Alternative Rock", "Egyptian rock band formed in 2003",
            "booking@cairokee.com", "", "",
            "Cairo", "Cairo Governorate", "Egypt",
            "https://open.spotify.com/artist/...", "https://youtube.com/@cairokee", "",
            "Berlin",
        ],
        [
            "Mashrou' Leila", "Indie Rock", "Lebanese indie band",
            "info@mashrouleila.com", "", "",
            "Beirut", "Beirut", "Lebanon",
            "https://open.spotify.com/artist/...", "", "",
            "London",
        ],
        [
            "Wegz", "Trap / Rap", "Egyptian rapper and singer",
            "", "", "",
            "Cairo", "", "Egypt",
            "https://open.spotify.com/artist/...", "https://instagram.com/wegz", "",
            "Dubai",
        ],
    ]
    for r, row_data in enumerate(examples, start=2):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(r, c, val)
            cell.font = EXAMPLE_FONT

    _auto_width(ws, min_width=18)

    # Instructions sheet
    ins = wb.create_sheet("Instructions")
    instructions = [
        ("Groovon — Artist List Template", Font(name="Calibri", bold=True, size=16, color="6C5CE7")),
        ("", None),
        ("Required Columns:", Font(name="Calibri", bold=True, size=12)),
        ("• Artist — Name of the artist (REQUIRED)", None),
        ("", None),
        ("Optional Columns:", Font(name="Calibri", bold=True, size=12)),
        ("• Genre — e.g., Rock, Pop, Hip-Hop, Electronic", None),
        ("• Bio — Short artist biography", None),
        ("• Email 1/2/3 — Contact emails", None),
        ("• Locale City/State/Country — Where the artist is based", None),
        ("• URL 1/2/3 — Spotify, YouTube, Instagram links", None),
        ("• List — The city where the artist is performing", None),
        ("", None),
        ("Notes:", Font(name="Calibri", bold=True, size=12)),
        ("• The sheet name should be 'Collection cleaned' (default)", None),
        ("• Row 1 = Headers, data starts from Row 2", None),
        ("• Only the 'Artist' column is required, the rest will be enriched by the pipeline", None),
        ("• If you only have artist names, just fill the Artist column!", None),
    ]
    for r, (text, font) in enumerate(instructions, start=1):
        cell = ins.cell(r, 1, text)
        if font:
            cell.font = font
        else:
            cell.font = Font(name="Calibri", size=11, color="444444")
    ins.column_dimensions["A"].width = 80

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Scraper JSON Template ───────────────────────────────────────────
def generate_scraper_template() -> bytes:
    """
    Returns a sample JSON file for ScraperSource.
    """
    import json
    sample = [
        {
            "title": "Indie Night — Mashrou' Leila Live",
            "city": "Berlin",
            "venue_name": "Lido",
            "date": "2025-07-15",
            "time": "21:00",
            "genre": "Indie Rock",
            "price": "EUR 25",
            "description": "Live concert featuring Mashrou' Leila",
            "image_url": "",
            "ticket_url": "https://example.com/tickets/123",
            "event_url": "https://example.com/events/mashrou-leila",
            "source": "custom_scraper",
            "source_id": "evt-001",
            "artists": ["Mashrou' Leila"],
            "artist_links": {
                "Mashrou' Leila": "https://open.spotify.com/artist/..."
            },
        },
        {
            "title": "Cairo Jazz Night",
            "city": "Cairo",
            "venue_name": "Cairo Jazz Club",
            "date": "2025-08-01",
            "time": "22:00",
            "genre": "Jazz",
            "price": "",
            "description": "",
            "image_url": "",
            "ticket_url": "",
            "event_url": "",
            "source": "custom_scraper",
            "source_id": "evt-002",
            "artists": [],
            "artist_links": {},
        },
    ]
    return json.dumps(sample, indent=2, ensure_ascii=False).encode("utf-8")
