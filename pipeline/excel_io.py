"""
pipeline/excel_io.py  —  Excel read / write helpers + confidence scoring.
Extracted from process_david_excel.py (pure extraction, zero logic changes).
"""

import json, logging, openpyxl
from openpyxl.styles import Font, PatternFill

from pipeline.config import (
    normalize, sanitize_event_type, validate_genre, best_genre,
)

log = logging.getLogger(__name__)

# Legacy constant / alias so the orchestrator's import resolves
HEADER_ROW = 3

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Email Ranking  (Item 5)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_EMAIL_PRIORITY = {"booking": 0, "management": 1, "press": 2, "general": 3}


def rank_emails(emails, email_labels=None):
    """Sort emails by label priority: booking > management > press > general > unlabelled.

    Parameters
    ----------
    emails : list[str]
    email_labels : dict[str, str] | None   – mapping email → label

    Returns
    -------
    list[str]  – emails sorted best-first
    """
    if not emails:
        return []
    labels = email_labels or {}

    def _key(em):
        lbl = labels.get(em, "unlabelled")
        return _EMAIL_PRIORITY.get(lbl, 99)

    return sorted(emails, key=_key)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Profile Score / Tier  (Item 6)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def calc_profile_score(profile):
    """Return (score_0_100, tier) for a merged artist profile.

    Scoring rubric (max ≈ 100):
      - Each platform link:    +5  (max 50 for 10 platforms)
      - Has bio:               +10
      - Has photo:             +5
      - Has genre (not fallback):  +5
      - Has at least 1 email:  +5
      - Has locale city:       +5
      - Verified on streaming: +10  (spotify OR deezer present)
      - Has upcoming events:   +5
      - Has top songs:         +5

    Tier:
      A  ≥ 70
      B  ≥ 40
      C  < 40
    """
    s = 0
    platforms = profile.get("platforms", {})
    s += min(len(platforms) * 5, 50)
    if profile.get("bio"):
        s += 10
    if profile.get("photo"):
        s += 5
    genre = profile.get("genre", "Don't Box Me!")
    if genre and genre != "Don't Box Me!":
        s += 5
    if profile.get("emails"):
        s += 5
    locale = profile.get("locale", {})
    if locale.get("city") or locale.get("active_city"):
        s += 5
    if "spotify" in platforms or "deezer" in platforms:
        s += 10
    if profile.get("upcoming_events") or profile.get("is_active"):
        s += 5
    if profile.get("top_songs"):
        s += 5

    score = min(s, 100)
    if score >= 70:
        tier = "A"
    elif score >= 40:
        tier = "B"
    else:
        tier = "C"
    return score, tier


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Confidence Scoring
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def calc_confidence(vf, enrich):
    """Calculate confidence: HIGH/MEDIUM/LOW based on verification and enrichment quality."""
    score = 0
    if vf:
        score += len(vf.get("sources", [])) * 2  # 2 points per platform
    if enrich:
        if enrich.get("bio") and len(str(enrich.get("bio", ""))) > 30:
            score += 2
        if enrich.get("locale_city"):
            score += 1
        if enrich.get("url1"):
            score += 1
        if enrich.get("email1"):
            score += 1
    if score >= 7:
        return "HIGH"
    elif score >= 4:
        return "MEDIUM"
    else:
        return "LOW"


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Input Validation  (Item 9)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def validate_inputs(path, sheet_name="base info"):
    """Pre-flight check on the input Excel file.

    Raises ValueError with a descriptive message on failure.
    Returns the number of data rows found on success.
    """
    import os
    if not os.path.isfile(path):
        raise ValueError(f"Input file not found: {path}")
    ext = os.path.splitext(path)[1].lower()
    if ext not in (".xlsx", ".xlsm"):
        raise ValueError(f"Unsupported file format '{ext}'. Expected .xlsx or .xlsm")
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot open workbook: {exc}") from exc
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    data_rows = sum(1 for r in range(4, ws.max_row + 1) if ws.cell(r, 4).value)
    wb.close()
    if data_rows == 0:
        raise ValueError("No data rows found (column D is empty from row 4 onwards).")
    log.info(f"   ✅ Input validated: {data_rows} data rows in sheet '{sheet_name}'")
    return data_rows


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Read Excel
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def read_excel(path, sheet_name="base info"):
    """Read event rows from an Excel workbook.
    Returns (workbook, worksheet, events_list).
    """
    wb = openpyxl.load_workbook(path)
    ws = wb[sheet_name]

    events = []
    for r in range(4, ws.max_row + 1):
        city = ws.cell(r, 2).value
        venue = ws.cell(r, 3).value
        title = ws.cell(r, 4).value
        if not title:
            continue
        events.append({
            "row": r,
            "city": str(city or "").strip(),
            "venue": str(venue or "").strip(),
            "title": str(title).strip(),
        })
    return wb, ws, events

# Legacy alias
read_events = read_excel

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Build Output Rows  (handles multi-artist row splitting)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_output_rows(classified):
    """Transform classified events into flat output rows (one row per artist)."""
    output_rows = []

    for e in classified:
        artists = e.get("artists", [])

        if e.get("delete"):
            output_rows.append({
                "id": "DELETE", "city": e["city"], "venue": e["venue"], "title": e["title"],
                "artist": None, "genre": e.get("genre"), "event_type": sanitize_event_type(e.get("event_type")),
                "delete": True, "is_event": True, "event_bio": e.get("event_bio")
            })
        elif not artists or e.get("is_event"):
            output_rows.append({
                "id": None, "city": e["city"], "venue": e["venue"], "title": e["title"],
                "artist": "Event", "genre": e.get("genre"), "event_type": sanitize_event_type(e.get("event_type")),
                "delete": False, "is_event": True, "event_bio": e.get("event_bio")
            })
        elif len(artists) == 1:
            output_rows.append({
                "id": None, "city": e["city"], "venue": e["venue"], "title": e["title"],
                "artist": artists[0], "genre": e.get("genre"), "event_type": sanitize_event_type(e.get("event_type")),
                "delete": False, "is_event": False
            })
        else:
            for artist_name in artists:
                output_rows.append({
                    "id": None, "city": e["city"], "venue": e["venue"], "title": e["title"],
                    "artist": artist_name, "genre": e.get("genre"), "event_type": sanitize_event_type(e.get("event_type")),
                    "delete": False, "is_event": False
                })

    return output_rows


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Write Output to Excel
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Cell fills
_DELETE_FILL = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")
_EVENT_FILL  = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_HIGH_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # green
_MED_FILL    = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")    # yellow
_LOW_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # red


def write_output(ws, output_rows, verify_cache, profile_cache, enrichment_cache,
                 wiki_cache, lastfm_cache, kg_cache, setlistfm_cache):
    """Write enriched output rows into the Excel worksheet.
    Returns confidence_counts dict {HIGH: n, MEDIUM: n, LOW: n}.
    """
    data_start_row = 4
    confidence_counts = {"HIGH": 0, "MEDIUM": 0, "LOW": 0}

    for out_idx, out in enumerate(output_rows):
        r = data_start_row + out_idx

        if out.get("delete"):
            ws.cell(r, 1).value = "DELETE"
            ws.cell(r, 1).fill = _DELETE_FILL

        ws.cell(r, 2).value = out["city"]
        ws.cell(r, 3).value = out["venue"]
        ws.cell(r, 4).value = out["title"]

        artist_name = out.get("artist")
        is_event = out.get("is_event", False)

        if artist_name and artist_name != "Event" and not is_event:
            norm = normalize(artist_name)
            vf = verify_cache.get(norm)
            profile = profile_cache.get(norm, {})
            enrich = enrichment_cache.get(norm, {})

            # Artist name — use profile/verified name
            ws.cell(r, 6).value = profile.get("name") or (vf["name"] if vf else artist_name)

            # Genre — priority: Spotify > AI > MusicBrainz
            genre = best_genre(out.get("genre", "Don't Box Me!"), vf)
            ws.cell(r, 7).value = genre

            # Event Type
            ws.cell(r, 8).value = out.get("event_type", "Live Music Performance")

            # Bio — from profile (merged from enrichment + scraped site)
            bio = profile.get("bio") or enrich.get("bio")
            if bio:
                ws.cell(r, 9).value = bio

            # Locale — from profile (merged from enrichment + MusicBrainz)
            locale = profile.get("locale", {})
            if locale.get("city"):
                ws.cell(r, 10).value = locale["city"]
            if locale.get("state"):
                ws.cell(r, 11).value = locale["state"]
            if locale.get("country"):
                ws.cell(r, 12).value = locale["country"]

            # URLs — from profile (merged + deduped + validated platforms)
            platforms = profile.get("platforms", {})
            url_priority = ["spotify", "deezer", "apple_music", "soundcloud", "youtube", "bandcamp",
                            "website", "instagram", "facebook", "twitter", "tiktok"]
            ordered_urls = []
            for p in url_priority:
                if p in platforms:
                    ordered_urls.append(platforms[p])
            for p, u in platforms.items():
                if u not in ordered_urls:
                    ordered_urls.append(u)
            for ui, url_val in enumerate(ordered_urls[:3]):
                ws.cell(r, 13 + ui).value = url_val

            # Emails — from profile (merged + validated)
            email_labels = profile.get("email_labels", {})
            for ei, em in enumerate(profile.get("emails", [])[:3]):
                label = email_labels.get(em)
                cell_val = f"{em} [{label}]" if label and label != "general" else em
                ws.cell(r, 16 + ei).value = cell_val

            # Confidence (column 19):
            # Prefer unified quality confidence if already computed in engine.
            conf = str(profile.get("confidence", "")).upper().strip()
            if conf not in {"HIGH", "MEDIUM", "LOW"}:
                conf = calc_confidence(vf, enrich)
                # Legacy boost logic kept as fallback.
                if conf == "MEDIUM":
                    if len(platforms) >= 4:
                        conf = "HIGH"
                    elif norm in wiki_cache or norm in lastfm_cache:
                        conf = "HIGH"
                    elif norm in kg_cache and kg_cache[norm].get("is_musician"):
                        conf = "HIGH"
                    elif norm in setlistfm_cache and setlistfm_cache[norm].get("total_setlists", 0) > 5:
                        conf = "HIGH"
                    elif profile.get("is_active"):
                        conf = "HIGH"
            confidence_counts[conf] = confidence_counts.get(conf, 0) + 1
            ws.cell(r, 19).value = conf
            if conf == "HIGH":
                ws.cell(r, 19).fill = _HIGH_FILL
            elif conf == "MEDIUM":
                ws.cell(r, 19).fill = _MED_FILL
            else:
                ws.cell(r, 19).fill = _LOW_FILL
        else:
            ws.cell(r, 6).value = "Event" if is_event else artist_name
            ws.cell(r, 7).value = validate_genre(out.get("genre", "Don't Box Me!"))
            ws.cell(r, 8).value = out.get("event_type", "Other")
            event_bio = out.get("event_bio")
            if event_bio and is_event:
                ws.cell(r, 9).value = event_bio
            if is_event and not out.get("delete"):
                ws.cell(r, 6).fill = _EVENT_FILL

    # Header for Confidence column
    ws.cell(3, 19).value = "Confidence"
    ws.cell(3, 19).font = Font(bold=True)

    return confidence_counts


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
#  Dump Rich Profiles to JSON
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def dump_profiles_json(path, profile_cache, verify_cache, enrichment_cache,
                       wiki_cache, lastfm_cache, kg_cache, setlistfm_cache):
    """Dump rich profiles with confidence scores to a JSON file."""
    rich_profiles = []
    for norm_name, profile in profile_cache.items():
        out = dict(profile)
        conf = str(out.get("confidence", "")).upper().strip()
        if conf not in {"HIGH", "MEDIUM", "LOW"}:
            conf = "LOW"
            if norm_name in verify_cache:
                enrich = enrichment_cache.get(norm_name, {})
                conf = calc_confidence(verify_cache.get(norm_name), enrich)
                if conf == "MEDIUM":
                    if len(out.get("platforms", {})) >= 4 or norm_name in wiki_cache or norm_name in lastfm_cache:
                        conf = "HIGH"
                    elif norm_name in kg_cache and kg_cache[norm_name].get("is_musician"):
                        conf = "HIGH"
                    elif norm_name in setlistfm_cache and setlistfm_cache[norm_name].get("total_setlists", 0) > 5:
                        conf = "HIGH"

        out["confidence"] = conf
        if "profile_score" not in out or "profile_tier" not in out:
            score, tier = calc_profile_score(out)
            out.setdefault("profile_score", score)
            out.setdefault("profile_tier", tier)

        rich_profiles.append(out)

    with open(path, "w", encoding="utf-8") as f:
        json.dump(rich_profiles, f, indent=2, ensure_ascii=False)

    return rich_profiles
