"""
sources.excel_source — Read events from David's Excel format.

Wraps the existing openpyxl reading logic from process_david_excel.py
(lines 106-122) behind the EventSource interface.

Config keys:
    path       (str)  – path to the .xlsx file (required)
    sheet      (str)  – worksheet name (default: "base info")
    start_row  (int)  – first data row (default: 4)
    col_city   (int)  – column index for city (default: 2)
    col_venue  (int)  – column index for venue (default: 3)
    col_title  (int)  – column index for title/event string (default: 4)
"""

from __future__ import annotations

import logging
import os
from typing import Any

from sources.base import EventSource
from pipeline.event_model import make_event

log = logging.getLogger(__name__)


class ExcelSource(EventSource):
    """Read events from a David-format Excel workbook."""

    name = "excel"

    def __init__(self, *, config: dict[str, Any] | None = None):
        super().__init__(config=config)
        self._path: str = self.config.get("path", "")
        self._sheet: str = self.config.get("sheet", "base info")
        self._start_row: int = self.config.get("start_row", 4)
        self._col_city: int = self.config.get("col_city", 2)
        self._col_venue: int = self.config.get("col_venue", 3)
        self._col_title: int = self.config.get("col_title", 4)

    # ── interface ───────────────────────────────────────────────────
    def validate(self) -> bool:
        if not self._path or not os.path.isfile(self._path):
            log.error(
                "ExcelSource: file not found → %s (cwd=%s, dir_exists=%s)",
                self._path,
                os.getcwd(),
                os.path.isdir(os.path.dirname(self._path)) if self._path else False,
            )
            return False
        return True

    def fetch(self) -> list[dict]:
        import openpyxl

        log.info(f"📖 ExcelSource: reading {self._path!r} (sheet={self._sheet!r})")
        wb = openpyxl.load_workbook(self._path)
        ws = wb[self._sheet]

        events: list[dict] = []
        for r in range(self._start_row, ws.max_row + 1):
            city = ws.cell(r, self._col_city).value
            venue = ws.cell(r, self._col_venue).value
            title = ws.cell(r, self._col_title).value
            if not title:
                continue
            events.append(
                make_event(
                    city=str(city or ""),
                    venue=str(venue or ""),
                    title=str(title),
                    row=r,
                    source="excel",
                )
            )

        log.info(f"   {len(events)} events read from Excel")
        return events
