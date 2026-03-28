"""
sources.artist_list_source — Read pre-extracted artist data from Excel.

Handles the format David/Yasser produce, where artists are already
extracted with genre, bio, emails, locale, and URLs.

Auto-detects sheets: "Collection cleaned" → "Cleaned list" → first
sheet with an "Artist" column header.

Config keys:
    path       (str)  – path to the .xlsx file (required)
    artists    (list[str]) – optional inline artist names (one per item)
    sheet      (str)  – worksheet name (auto-detect if omitted)
    start_row  (int)  – first data row; default: 2 (row 1 = headers)
"""

from __future__ import annotations

import logging
import os
from typing import Any

from sources.base import EventSource
from pipeline.event_model import make_event

log = logging.getLogger(__name__)

# Sheets to try, in order, when auto-detecting
_PREFERRED_SHEETS = ["Collection cleaned", "Cleaned list"]

# Header names (case-insensitive) → canonical key
_HEADER_MAP: dict[str, str] = {
    "artist": "artist",
    "genre": "genre",
    "bio": "bio",
    "email 1": "email_1",
    "email 2": "email_2",
    "email 3": "email_3",
    "locale city": "locale_city",
    "local state": "locale_state",
    "local country": "locale_country",
    "url 1": "url_1",
    "url2": "url_2",
    "url 2": "url_2",
    "url3": "url_3",
    "url 3": "url_3",
    "performing city": "performing_city",
    "list": "performing_city",       # "Collection cleaned" uses "List" for city
    "dbs id": "dbs_id",
}


class ArtistListSource(EventSource):
    """Read pre-extracted artist lists from David/Yasser Excel workbooks."""

    name = "artist_list"
    source_type = "artist_list"

    def __init__(self, *, config: dict[str, Any] | None = None):
        super().__init__(config=config)
        self._path: str = self.config.get("path", "")
        self._artists: list[str] = [str(a).strip() for a in self.config.get("artists", []) if str(a).strip()]
        self._sheet: str | None = self.config.get("sheet")
        self._start_row: int = self.config.get("start_row", 2)
        self._city: str = str(self.config.get("city", "")).strip()
        self._genre: str = str(self.config.get("genre", "")).strip()

    # ── interface ───────────────────────────────────────────────────
    def validate(self) -> bool:
        if self._artists:
            return True
        if self._path and os.path.isfile(self._path):
            return True
        log.error(f"ArtistListSource: file not found → {self._path}")
        return False

    def fetch(self) -> list[dict]:
        import openpyxl

        if self._artists and not self._path:
            events: list[dict] = []
            for idx, artist_name in enumerate(self._artists, start=1):
                event = make_event(
                    city=self._city,
                    venue="",
                    title=artist_name,
                    row=idx,
                    genre=self._genre,
                    source="artist_list",
                )
                event["artists"] = [artist_name]
                event["genre"] = self._genre or "Don't Box Me!"
                event["event_type"] = "Pre-extracted"
                event["delete"] = False
                event["is_event"] = False
                event["has_more"] = False
                event["event_bio"] = None
                event["_seed"] = {
                    "bio": "",
                    "emails": [],
                    "urls": [],
                    "locale_city": "",
                    "locale_state": "",
                    "locale_country": "",
                    "dbs_id": "",
                }
                events.append(event)
            log.info(f"   {len(events)} artists read from inline list")
            return events

        log.info(f"📖 ArtistListSource: reading {self._path!r}")
        wb = openpyxl.load_workbook(self._path, data_only=True)

        # ── pick worksheet ──────────────────────────────────────────
        ws = None
        sheet_name = self._sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Sheet {sheet_name!r} not found. "
                    f"Available: {wb.sheetnames}"
                )
            ws = wb[sheet_name]
        else:
            # auto-detect
            for candidate in _PREFERRED_SHEETS:
                if candidate in wb.sheetnames:
                    ws = wb[candidate]
                    sheet_name = candidate
                    break
            if ws is None:
                # fallback: first sheet with an "Artist" column
                for sn in wb.sheetnames:
                    s = wb[sn]
                    for c in range(1, min(s.max_column or 1, 30) + 1):
                        val = s.cell(1, c).value
                        if val and str(val).strip().lower() == "artist":
                            ws = s
                            sheet_name = sn
                            break
                    if ws:
                        break
            if ws is None:
                raise ValueError(
                    f"Cannot auto-detect artist sheet. "
                    f"Available sheets: {wb.sheetnames}"
                )

        log.info(f"   Using sheet: {sheet_name!r}")

        # ── detect column mapping from header row ───────────────────
        col_map: dict[str, int] = {}  # canonical_key → column index
        for c in range(1, (ws.max_column or 1) + 1):
            raw = ws.cell(1, c).value
            if raw is None:
                continue
            header = str(raw).strip().lower()
            if header in _HEADER_MAP:
                col_map[_HEADER_MAP[header]] = c

        if "artist" not in col_map:
            raise ValueError(
                f"Sheet {sheet_name!r} has no 'Artist' column. "
                f"Headers: {[ws.cell(1, c).value for c in range(1, (ws.max_column or 1) + 1)]}"
            )

        log.info(f"   Columns detected: {list(col_map.keys())}")

        # ── read rows ───────────────────────────────────────────────
        def _cell(row: int, key: str) -> str:
            col = col_map.get(key)
            if col is None:
                return ""
            val = ws.cell(row, col).value
            return str(val).strip() if val is not None else ""

        events: list[dict] = []
        for r in range(self._start_row, (ws.max_row or 1) + 1):
            artist_name = _cell(r, "artist")
            if not artist_name:
                continue

            city = _cell(r, "performing_city") or _cell(r, "locale_city")
            genre = _cell(r, "genre")
            bio = _cell(r, "bio")

            # Collect emails
            emails = [
                e for e in [_cell(r, "email_1"), _cell(r, "email_2"), _cell(r, "email_3")]
                if e
            ]

            # Collect URLs
            urls = [
                u for u in [_cell(r, "url_1"), _cell(r, "url_2"), _cell(r, "url_3")]
                if u
            ]

            # Locale
            locale_city = _cell(r, "locale_city")
            locale_state = _cell(r, "locale_state")
            locale_country = _cell(r, "locale_country")

            # Build an event dict that mimics post-classification output
            event = make_event(
                city=city,
                venue="",
                title=artist_name,
                row=r,
                genre=genre,
                source="artist_list",
            )
            # Add pre-classified fields so engine can skip Phase 1
            event["artists"] = [artist_name]
            event["genre"] = genre or "Don't Box Me!"
            event["event_type"] = "Pre-extracted"
            event["delete"] = False
            event["is_event"] = False
            event["has_more"] = False
            event["event_bio"] = bio

            # Carry forward the extra data for enrichment seeding
            event["_seed"] = {
                "bio": bio,
                "emails": emails,
                "urls": urls,
                "locale_city": locale_city,
                "locale_state": locale_state,
                "locale_country": locale_country,
                "dbs_id": _cell(r, "dbs_id"),
            }

            events.append(event)

        log.info(f"   {len(events)} artists read from Excel")
        return events
