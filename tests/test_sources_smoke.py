import tempfile
import unittest
from pathlib import Path

import openpyxl

from sources.api_source import APISource
from sources.artist_list_source import ArtistListSource
from sources.excel_source import ExcelSource


class SourceSmokeTests(unittest.TestCase):
    def test_excel_source_reads_events(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "events.xlsx"
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "base info"
            ws.cell(4, 2).value = "Berlin"
            ws.cell(4, 3).value = "Lido"
            ws.cell(4, 4).value = "Indie Night"
            ws.cell(5, 2).value = ""
            ws.cell(5, 3).value = ""
            ws.cell(5, 4).value = ""
            wb.save(path)

            source = ExcelSource(config={"path": str(path)})
            self.assertTrue(source.validate())
            events = source.fetch()
            self.assertEqual(len(events), 1)
            self.assertEqual(events[0]["city"], "Berlin")
            self.assertEqual(events[0]["venue"], "Lido")
            self.assertEqual(events[0]["title"], "Indie Night")

    def test_api_source_validate_requires_city(self):
        self.assertFalse(APISource(config={}).validate())
        self.assertTrue(APISource(config={"city": "Berlin"}).validate())

    def test_artist_list_source_inline_artists(self):
        source = ArtistListSource(
            config={
                "artists": ["Artist A", "Artist B"],
                "city": "Cairo",
                "genre": "Jazz",
            }
        )

        self.assertTrue(source.validate())
        events = source.fetch()
        self.assertEqual(len(events), 2)
        self.assertEqual(events[0]["artists"], ["Artist A"])
        self.assertEqual(events[0]["city"], "Cairo")
        self.assertEqual(events[0]["genre"], "Jazz")
        self.assertEqual(events[0]["source"], "artist_list")


if __name__ == "__main__":
    unittest.main()
