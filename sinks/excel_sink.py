"""
sinks.excel_sink — Write enriched pipeline output to an Excel workbook.

This is a thin adapter: the heavy formatting/cell-writing logic remains
in `pipeline.excel_io.write_output`.  The sink's job is to call it with
the right arguments and return the output path.

Config keys:
    output_path  (str) – path to write the .xlsx file (required)
    input_path   (str) – path to the source .xlsx (used as template if style
                         needs to be preserved; optional)
"""

from __future__ import annotations

import logging
import os
from typing import Any

from sinks.base import OutputSink

log = logging.getLogger(__name__)


class ExcelSink(OutputSink):
    """Write enriched data to David-format Excel."""

    name = "excel"

    def __init__(self, *, config: dict[str, Any] | None = None):
        super().__init__(config=config)
        self._output_path: str = self.config.get("output_path", "")
        self._input_path: str = self.config.get("input_path", "")

    def validate(self) -> bool:
        if not self._output_path:
            log.error("ExcelSink: 'output_path' config is required")
            return False
        return True

    def write(
        self,
        *,
        classified: list[dict],
        verify_cache: dict,
        enrichment_cache: dict,
        profile_cache: dict,
        output_rows: list[dict],
    ) -> str | None:
        """Delegate to the existing excel_io.write_output function."""
        import openpyxl
        from pipeline.excel_io import write_output

        try:
            wb = None
            if self._input_path and os.path.isfile(self._input_path):
                wb = openpyxl.load_workbook(self._input_path)
                ws = wb["base info"] if "base info" in wb.sheetnames else wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "base info"

            write_output(
                ws=ws,
                output_rows=output_rows,
                verify_cache=verify_cache,
                profile_cache=profile_cache,
                enrichment_cache=enrichment_cache,
                wiki_cache={},
                lastfm_cache={},
                kg_cache={},
                setlistfm_cache={},
            )
            os.makedirs(os.path.dirname(self._output_path) or ".", exist_ok=True)
            wb.save(self._output_path)
            log.info(f"📄 ExcelSink: wrote {self._output_path}")
            return self._output_path
        except Exception as exc:
            log.error(f"ExcelSink: write failed → {exc}")
            return None
